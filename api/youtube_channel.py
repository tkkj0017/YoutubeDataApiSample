import openpyxl
import pytz
from openpyxl.styles.borders import Border, Side
from googleapiclient.discovery import build
from _datetime import datetime, date, time
from dateutil import relativedelta

# 要件
# OUTPUT... チャンネル名(snippet.title)、URL(URL + id)、登録者数(statistics.subscriberCount)
# ・検索キーワード...株式会社
# ・チャンネル登録者100人以上(channels apiでstatistics.subscriberCountを取得しフィルター)
# ・全投稿数10本以上(channels apiでstatistics.videoCountを取得しフィルター)
# ・1カ月以内投稿頻度4以上(search apiでビデオデータ取得後、フィルター)　

# APIキーの定義(自分のGoogleアカウントのGCPにて取得し設定)
API_KEY = "%APIキー%"
# youtube APIクライアント
youtube = build("youtube", "v3", developerKey=API_KEY)
# Excelテンプレート
excel_template = "sample.xlsx"

# 検索条件↓
# 検索キーワード
search_word = "株式会社"
# 最低チャンネル登録者数(〜人以上)
fl_subscriber_count = 100
# 動画投稿数
fl_video_count = 10
# 直近1ヶ月以内の最低投稿数(~本以上)
fl_video_count_a_month = 4


# 実行時より1ヶ月前の日付を取得
def get_a_month_ago_date():
    today = date.today()
    a_month_ago_date = today + relativedelta.relativedelta(months=-1)
    dt_native = datetime.combine(a_month_ago_date, time())
    a_month_ago_date_tz = pytz.timezone('Asia/Tokyo').localize(dt_native)
    iso_month_ago_date = datetime.strftime(a_month_ago_date_tz, '%Y-%m-%dT%H:%M:%S.%fZ')
    return iso_month_ago_date


# キーワードからチャンネルを検索
def get_channel_list():
    # nextPagetokenがなくなるまで処理を実行しリストに追加
    ch_list = []
    next_page_token = 'start'
    while next_page_token is not None:
        if next_page_token == 'start':
            search_response = youtube.search().list(
                q=search_word,
                part='snippet, id',
                maxResults=50,
                type='channel'
            ).execute()
        else:
            search_response = youtube.search().list(
                q=search_word,
                part='snippet, id',
                maxResults=50,
                pageToken=next_page_token,
                type='channel'
            ).execute()

        if 'nextPageToken' in search_response:
            next_page_token = search_response['nextPageToken']
        else:
            next_page_token = None

        for item in search_response['items']:
            ch_list.append(item)

    print("=== チャンネル検索完了 ===")
    return ch_list


# チャンネルをフィルタリングし、出力リストを洗い出す
def filter_channels(ch_list):
    # チャンネルリスト取得後のフィルター処理
    output_data = []
    for channel in ch_list:
        print(str(channel['snippet']['title']) + "の情報取得")

        # チャンネル詳細情報取得
        channel_response = youtube.channels().list(
            part='id, statistics',
            id=channel['id']['channelId']
        ).execute()

        # チャンネル登録者数が確認できなければ、次のループ(出力候補から外す)
        if not channel_response['items'][0]['statistics'].get('subscriberCount'):
            continue

        # チャンネル登録者数が◯人以上いなければ、次のループ(出力候補から外す)
        if not int(channel_response['items'][0]['statistics']['subscriberCount']) >= fl_subscriber_count:
            continue

        # アカウントの動画投稿数が◯本以上でなければ、次のループ(出力候補から外す)
        if not int(channel_response['items'][0]['statistics']['videoCount']) >= fl_video_count:
            continue

        # TODO 「1ヶ月以内投稿数4以上」を調べるためにリクエストを投げるが、ここでAPI使用量の上限に達してしまうため現状保留
        # TODO 対応策... チャンネル登録者数の条件を5000~10000人以上にする(条件厳しくしてリクエスト数を減らす)。API使用量の増枠を申請するetc...
        # # チャンネルIDで動画検索
        # iso_published_after = get_a_month_ago_date()
        # video_response = youtube.search().list(
        #     part='snippet',
        #     channelId=channel['id']['channelId'],
        #     publishedAfter=iso_published_after,
        #     maxResults=10,
        #     type='video'
        #     ).execute()
        #
        # #  1ヶ月以内の投稿数が◯回より少なければ、次のループ(出力候補から外す)
        # if len(video_response['items']) < fl_video_count_a_month:
        #     continue

        # 登録者数のキーを独自で追加
        channel['subscriberCount'] = channel_response['items'][0]['statistics']['subscriberCount']
        # 出力対象リストに追加
        output_data.append(channel)

    print("出力データ数: " + str(len(output_data)))
    return output_data


# 取得したデータをExcelに出力
def save_excel_file(output_data):
    # Excelファイルダウンロード
    wb = openpyxl.load_workbook(excel_template)
    ws = wb['channels']
    border = Border(top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'),
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000')
                    )
    row = 2
    for item in output_data:
        ws.cell(row, 1).value = item['snippet']['title']
        ws.cell(row, 2).value = item['subscriberCount']
        ws.cell(row, 3).value = "https://www.youtube.com/channel/" + item['id']['channelId']

        ws.cell(row, 1).border = border
        ws.cell(row, 2).border = border
        ws.cell(row, 3).border = border
        row += 1
    dt_now = datetime.now()
    wb.save("チャンネルリスト抽出結果" + dt_now.strftime('_%Y%m%d_%H%M') + '.xlsx')


channel_list = get_channel_list()
output = filter_channels(channel_list)
save_excel_file(output)
